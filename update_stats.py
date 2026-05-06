"""
주가지수 월간 통계 업데이트 스크립트
=====================================
매월 1일 GitHub Actions에서 자동 실행됨.

실행 내용:
1. yfinance에서 S&P 500, NASDAQ 100, KOSPI 월간 가격 다운로드
2. 월간 수익률 계산
3. 통계 (평균, σ, Sharpe, VaR/CVaR, 첨도/왜도) 계산
4. 계절성 (월별 평균) 계산
5. 상관관계 매트릭스 계산
6. docs/index.html 생성 (사이드바 아티팩트와 동일 디자인)
7. docs/data.xlsx 생성 (Excel 파일)
"""

import json
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ============================================================
# CONFIG
# ============================================================
TICKERS = {
    "S&P 500": "^GSPC",
    "NASDAQ 100": "^NDX",
    "KOSPI": "^KS11",
}

START_DATE = "2005-12-01"  # 2006년 1월 첫 수익률 계산을 위해 한 달 전부터
ROOT = Path(__file__).parent
DOCS = ROOT / "docs"
TEMPLATE = ROOT / "templates" / "index_template.html"

# Fed/BOK 연말 금리는 자동 수집이 어려워 수동 관리. 매년 12월에만 업데이트하면 됨.
FED_BOK_RATES = {
    2006: (5.25, 4.50), 2007: (4.25, 5.00), 2008: (0.25, 3.00),
    2009: (0.25, 2.00), 2010: (0.25, 2.50), 2011: (0.25, 3.25),
    2012: (0.25, 2.75), 2013: (0.25, 2.50), 2014: (0.25, 2.00),
    2015: (0.50, 1.50), 2016: (0.75, 1.25), 2017: (1.50, 1.50),
    2018: (2.50, 1.75), 2019: (1.75, 1.25), 2020: (0.25, 0.50),
    2021: (0.25, 1.00), 2022: (4.50, 3.25), 2023: (5.50, 3.50),
    2024: (4.50, 3.00), 2025: (3.75, 2.50), 2026: (3.75, 2.50),
}

# 연도별 주요 이벤트 (수동 관리 - 새해 1월 한 번만 추가)
YEAR_EVENTS = {
    2006: "Fed 5회 인상 사이클 정점",
    2007: "서브프라임 위기 시작 / KOSPI 2000 첫 돌파",
    2008: "글로벌 금융위기 (리먼 사태)",
    2009: "QE1 시작 / 위기 후 반등",
    2010: "QE2",
    2011: "유럽 재정위기 / 美 신용등급 강등",
    2012: "QE3",
    2013: "美 강세장 / 한국 박스피",
    2014: "QE 종료 / 유가 급락",
    2015: "Fed 첫 인상 (12월)",
    2016: "Brexit / 트럼프 당선",
    2017: "동기화 글로벌 성장 / 코스피 2500 돌파",
    2018: "美中 무역전쟁 / Q4 폭락",
    2019: "Fed 보험성 인하",
    2020: "COVID-19 / 무제한 QE",
    2021: "백신 랠리 / 코스피 3000 돌파",
    2022: "Fed 공격적 인상 / 우크라 전쟁",
    2023: "AI 랠리 (엔비디아) / 디스인플레이션",
    2024: "美 AI 빅테크 독주 / 韓 12·3 계엄 충격",
    2025: "韓 밸류업+AI 반도체 호황 / G20 1위",
    2026: "이란 전쟁 / 韓 코스피 6000 돌파",
}


# ============================================================
# 데이터 수집
# ============================================================
def fetch_monthly_data() -> dict:
    """yfinance에서 월간 종가 다운로드"""
    today = datetime.today().strftime("%Y-%m-%d")
    data = {}
    for name, ticker in TICKERS.items():
        df = yf.download(
            ticker, start=START_DATE, end=today,
            interval="1mo", progress=False, auto_adjust=False,
        )
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
        data[name] = df
        print(f"[OK] {name}: {len(df)}개월 ({df.index[0].date()} ~ {df.index[-1].date()})")
    return data


def build_returns_df(raw: dict) -> pd.DataFrame:
    """월간 종가 → 월간 수익률(%) DataFrame"""
    closes = pd.DataFrame({name: df["Close"] for name, df in raw.items()})
    returns = closes.pct_change() * 100
    return returns.loc["2006-01-01":]


# ============================================================
# 통계 계산
# ============================================================
def compute_stats(series: pd.Series) -> dict:
    s = series.dropna()
    n = len(s)
    return {
        "n": n,
        "mean": float(s.mean()),
        "median": float(s.median()),
        "std": float(s.std()),
        "ann_vol": float(s.std() * np.sqrt(12)),
        "min": float(s.min()),
        "min_month": s.idxmin().strftime("%Y-%m"),
        "max": float(s.max()),
        "max_month": s.idxmax().strftime("%Y-%m"),
        "skew": float(s.skew()),
        "kurtosis": float(s.kurtosis()),
        "hit_rate": float((s > 0).sum() / n * 100),
        "pos_count": int((s > 0).sum()),
        "neg_count": int((s < 0).sum()),
        "var_5": float(s.quantile(0.05)),
        "cvar_5": float(s[s <= s.quantile(0.05)].mean()),
        "cagr": float(((1 + s / 100).prod() ** (12 / n) - 1) * 100),
        "sharpe": float((s.mean() / s.std()) * np.sqrt(12)),
    }


def compute_seasonality(returns: pd.DataFrame) -> dict:
    out = {}
    for col in returns.columns:
        s = returns[col].dropna()
        out[col] = {
            m: {
                "mean": float(s[s.index.month == m].mean()),
                "hit_rate": float((s[s.index.month == m] > 0).sum() / len(s[s.index.month == m]) * 100),
            }
            for m in range(1, 13)
        }
    return out


def compute_annual_returns(raw: dict) -> dict:
    """각 연도별 가격 변동률 계산 (Price Return)"""
    out = {}
    for name, df in raw.items():
        closes = df["Close"]
        # 연말 종가 기준
        yearly = closes.resample("YE").last()
        annual_pct = yearly.pct_change() * 100
        # 2026년은 YTD = 가장 최근 종가 / 2025-12-31 종가 - 1
        latest_year = datetime.today().year
        if latest_year == yearly.index[-1].year:
            ytd_base = yearly.iloc[-2]  # 작년 말
            ytd_latest = closes.iloc[-1]  # 가장 최근
            annual_pct.iloc[-1] = (ytd_latest / ytd_base - 1) * 100
        out[name] = {idx.year: float(val) for idx, val in annual_pct.dropna().items()}
    return out


# ============================================================
# Excel 생성
# ============================================================
def build_excel(returns: pd.DataFrame, stats: dict, seasonality: dict,
                corr: pd.DataFrame, annual: dict, out_path: Path) -> None:
    wb = Workbook()

    HEADER_FILL = PatternFill("solid", start_color="0C2340")
    HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(name="Arial", size=10)
    INPUT_FONT = Font(name="Arial", color="0000FF", size=10)
    BORDER = Border(*[Side(border_style="thin", color="BFBFBF")] * 4)
    POS_FILL = PatternFill("solid", start_color="E2EFDA")
    NEG_FILL = PatternFill("solid", start_color="FCE4D6")

    # === Sheet 1: 연간 수익률 & 금리 ===
    ws = wb.active
    ws.title = "연간 수익률&금리"
    ws["A1"] = f"주가지수 연간 수익률 & 금리 (2006~{datetime.today().year} YTD)"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="0C2340")
    ws.merge_cells("A1:G1")
    ws["A2"] = f"업데이트: {datetime.today().strftime('%Y-%m-%d')} | Yahoo Finance Price Return"
    ws["A2"].font = Font(name="Arial", italic=True, color="808080", size=9)
    ws.merge_cells("A2:G2")

    headers = ["연도", "S&P 500 (%)", "NASDAQ 100 (%)", "KOSPI (%)",
               "Fed Funds (연말)", "BOK (연말)", "주요 이벤트"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    years = sorted(set(annual["S&P 500"].keys()))
    for i, yr in enumerate(years, 5):
        sp = annual["S&P 500"].get(yr)
        nd = annual["NASDAQ 100"].get(yr)
        kp = annual["KOSPI"].get(yr)
        fed, bok = FED_BOK_RATES.get(yr, (None, None))
        ev = YEAR_EVENTS.get(yr, "")

        ws.cell(row=i, column=1, value=yr).number_format = "0"
        for col, val in zip([2, 3, 4], [sp, nd, kp]):
            cell = ws.cell(row=i, column=col, value=round(val, 2) if val is not None else None)
            cell.number_format = "0.00"
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="center")
            if val is not None:
                cell.fill = POS_FILL if val >= 0 else NEG_FILL
        ws.cell(row=i, column=5, value=fed).number_format = "0.00"
        ws.cell(row=i, column=6, value=bok).number_format = "0.00"
        ws.cell(row=i, column=7, value=ev)
        for c in range(1, 8):
            ws.cell(row=i, column=c).border = BORDER
            if c == 7:
                ws.cell(row=i, column=c).font = NORMAL_FONT

    # 통계 요약
    summary_row = 5 + len(years) + 2
    ws.cell(row=summary_row, column=1, value="통계 요약").font = Font(name="Arial", bold=True, color="0C2340")
    cols = ["S&P 500", "NASDAQ 100", "KOSPI"]
    for j, name in enumerate(cols):
        st = stats[name]
        rows = [
            ("CAGR (월간복리, 연환산 %)", st["cagr"]),
            ("월 평균 (%)", st["mean"]),
            ("연환산 변동성 (%)", st["ann_vol"]),
            ("Sharpe (RF=0)", st["sharpe"]),
            ("최저 월", f"{st['min_month']} ({st['min']:.2f}%)"),
            ("최고 월", f"{st['max_month']} ({st['max']:.2f}%)"),
        ]
        col_offset = 2 + j
        ws.cell(row=summary_row + 1, column=col_offset, value=name).font = HEADER_FONT
        ws.cell(row=summary_row + 1, column=col_offset).fill = HEADER_FILL
        for k, (lbl, val) in enumerate(rows):
            if j == 0:
                ws.cell(row=summary_row + 2 + k, column=1, value=lbl)
            cell = ws.cell(row=summary_row + 2 + k, column=col_offset)
            if isinstance(val, float):
                cell.value = round(val, 3)
                cell.number_format = "0.000"
            else:
                cell.value = val
            cell.alignment = Alignment(horizontal="center")

    for col, w in zip("ABCDEFG", [10, 14, 16, 14, 16, 14, 50]):
        ws.column_dimensions[col].width = w

    # === Sheet 2: 월간 수익률 전체 ===
    ws2 = wb.create_sheet("월간 수익률")
    ws2["A1"] = f"월간 수익률 (2006.1 ~ {returns.index[-1].strftime('%Y.%m')})"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="0C2340")
    ws2.merge_cells("A1:D1")

    headers2 = ["연월", "S&P 500 (%)", "NASDAQ 100 (%)", "KOSPI (%)"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (idx, row) in enumerate(returns.iterrows(), 4):
        ws2.cell(row=i, column=1, value=idx.strftime("%Y-%m")).alignment = Alignment(horizontal="center")
        for j, name in enumerate(cols):
            v = row[name]
            cell = ws2.cell(row=i, column=2 + j, value=round(v, 2) if pd.notna(v) else None)
            cell.number_format = "0.00"
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="center")
            if pd.notna(v):
                cell.fill = POS_FILL if v >= 0 else NEG_FILL

    ws2.freeze_panes = "A4"
    for col, w in zip("ABCD", [12, 14, 16, 14]):
        ws2.column_dimensions[col].width = w

    # === Sheet 3: 통계 ===
    ws3 = wb.create_sheet("통계")
    ws3["A1"] = "월간 통계 (전 기간)"
    ws3["A1"].font = Font(name="Arial", bold=True, size=14, color="0C2340")
    ws3.merge_cells("A1:D1")

    metrics = [
        ("관측치 (월)", "n", "0"),
        ("월 평균 (%)", "mean", "0.000"),
        ("중위값 (%)", "median", "0.000"),
        ("표준편차 (월,%)", "std", "0.000"),
        ("연환산 변동성 (%)", "ann_vol", "0.0"),
        ("최저 월 (%)", "min", "0.00"),
        ("최저 월 시점", "min_month", None),
        ("최고 월 (%)", "max", "0.00"),
        ("최고 월 시점", "max_month", None),
        ("왜도", "skew", "0.000"),
        ("첨도", "kurtosis", "0.00"),
        ("상승월 비율 (%)", "hit_rate", "0.0"),
        ("VaR 5% (%)", "var_5", "0.00"),
        ("CVaR 5% (%)", "cvar_5", "0.00"),
        ("CAGR (연환산 %)", "cagr", "0.00"),
        ("Sharpe-like", "sharpe", "0.000"),
    ]
    headers3 = ["지표", "S&P 500", "NASDAQ 100", "KOSPI"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=3, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (lbl, key, fmt) in enumerate(metrics, 4):
        ws3.cell(row=i, column=1, value=lbl).font = Font(name="Arial", bold=True, size=10)
        for j, name in enumerate(cols):
            cell = ws3.cell(row=i, column=2 + j, value=stats[name][key])
            cell.font = INPUT_FONT
            cell.alignment = Alignment(horizontal="center")
            if fmt:
                cell.number_format = fmt

    for col, w in zip("ABCD", [22, 14, 16, 14]):
        ws3.column_dimensions[col].width = w

    wb.save(out_path)
    print(f"[OK] Excel 저장: {out_path}")


# ============================================================
# HTML 생성
# ============================================================
def build_html(returns: pd.DataFrame, stats: dict, seasonality: dict,
               corr: pd.DataFrame, annual: dict, out_path: Path) -> None:
    template_text = TEMPLATE.read_text(encoding="utf-8")

    # 데이터 패키징
    payload = {
        "updated_at": datetime.today().strftime("%Y-%m-%d"),
        "period_start": "2006.1",
        "period_end": returns.index[-1].strftime("%Y.%m"),
        "n_months": len(returns),
        "stats": stats,
        "seasonality": seasonality,
        "correlation": corr.round(3).to_dict(),
        "annual": [
            [
                yr,
                round(annual["S&P 500"].get(yr), 2) if annual["S&P 500"].get(yr) is not None else None,
                round(annual["NASDAQ 100"].get(yr), 2) if annual["NASDAQ 100"].get(yr) is not None else None,
                round(annual["KOSPI"].get(yr), 2) if annual["KOSPI"].get(yr) is not None else None,
                FED_BOK_RATES.get(yr, (None, None))[0],
                FED_BOK_RATES.get(yr, (None, None))[1],
                YEAR_EVENTS.get(yr, ""),
            ]
            for yr in sorted(annual["S&P 500"].keys())
        ],
    }

    json_str = json.dumps(payload, ensure_ascii=False, indent=2)
    html = template_text.replace("__DATA_JSON__", json_str)
    out_path.write_text(html, encoding="utf-8")
    print(f"[OK] HTML 저장: {out_path}")


# ============================================================
# Main
# ============================================================
def main() -> None:
    print(f"\n{'=' * 60}")
    print(f"실행 시각: {datetime.today().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 60}\n")

    print("[1/5] 데이터 다운로드...")
    raw = fetch_monthly_data()

    print("\n[2/5] 월간 수익률 계산...")
    returns = build_returns_df(raw)
    print(f"      기간: {returns.index[0].date()} ~ {returns.index[-1].date()} ({len(returns)}개월)")

    print("\n[3/5] 통계 계산...")
    stats = {col: compute_stats(returns[col]) for col in returns.columns}
    seasonality = compute_seasonality(returns)
    corr = returns.corr()
    annual = compute_annual_returns(raw)

    for name in TICKERS.keys():
        s = stats[name]
        print(f"      {name}: CAGR {s['cagr']:.2f}% / σ_ann {s['ann_vol']:.1f}% / "
              f"Sharpe {s['sharpe']:.2f}")

    print("\n[4/5] Excel 생성...")
    DOCS.mkdir(exist_ok=True)
    build_excel(returns, stats, seasonality, corr, annual, DOCS / "data.xlsx")

    print("\n[5/5] HTML 생성...")
    build_html(returns, stats, seasonality, corr, annual, DOCS / "index.html")

    print(f"\n{'=' * 60}")
    print("✅ 완료!")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
